# -*- coding: utf-8 -*-
"""出荷連絡表 自動生成エンジン
仮納品書(Excel) + 商品マスタ(xlsm/xlsx/csv) -> 単価入り出荷連絡表(xlsx)

ルール（要件定義 v1 / 2026-06-18 確定分）:
- 入力は仮納品書。タブ=納品先ごと。1製品が複数先に分納 -> 製品×ロットで数量合算。
- 資材(「残資材」ブロック・容器/ポンプ等)は除外。有償サンプルは含む。
- 1ファイル=その日の出荷連絡表1枚（日付フィルタなし）。受領書ブロックは二重計上を避けて無視。
- 単価は商品マスタから品名の正規化突合で取得（NFKC+空白除去+小文字）。
- 数量帯: ロット欄の下限数値を閾値とし「数量以下で最大の閾値」を採用。
         最小閾値未満なら最小単価を採用しつつ「要確認」。
- 価格履歴: 同条件で複数あれば最新の価格更新日を採用。
- 名前衝突(正規化後に別商品CDが複数): 「要確認」。
- 単価が引けない/曖昧は「要確認」で別掲（黙って誤単価を出さない）。
"""
import re
import csv
import io
import unicodedata
import datetime
from collections import defaultdict, Counter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------- 正規化ユーティリティ ----------------
def normalize(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = re.sub(r"\s+", "", s)
    return s.lower()

def first_num(s):
    if s is None:
        return None
    s = unicodedata.normalize("NFKC", str(s))
    m = re.search(r"\d[\d,]*(?:\.\d+)?", s)
    return float(m.group(0).replace(",", "")) if m else None

def parse_date(s):
    if s is None:
        return None
    if isinstance(s, datetime.datetime):   # datetime は date より先に判定（date のサブクラスのため）
        return s.date()
    if isinstance(s, datetime.date):
        return s
    if isinstance(s, (int, float)) and 20000 <= s <= 80000:  # Excelシリアル値（1954〜2119頃）
        return (datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(s))).date()
    s = unicodedata.normalize("NFKC", str(s))
    m = re.search(r"(\d{4})[./\-](\d{1,2})[./\-](\d{1,2})", s)  # YYYY/M/D
    if m:
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    m = re.search(r"(\d{1,2})[./\-](\d{1,2})[./\-](\d{4})", s)  # M/D/YYYY（仮納品書の日付）
    if m:
        try:
            return datetime.date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
        except ValueError:
            return None
    return None

# ---------------- 資材判定 ----------------
SHIZAI_EXACT = {"容器", "ポンプ", "キャップ", "化粧箱", "6コ箱", "外装", "ラベル",
                "シール", "中栓", "スパチュラ", "袋", "箱", "台紙", "パンフ",
                "リーフレット", "説明書", "個箱", "内箱", "外箱",
                "パウチ", "シュリンク", "スポイド", "スポイト", "バーコードラベル", "副資材"}
_SHIZAI_NORM = {normalize(x) for x in SHIZAI_EXACT}
# 「○○ラベル」「○○パウチ」等、資材名で終わる行も資材とみなす（語尾一致）
SHIZAI_SUFFIX = ("バーコードラベル", "ラベル", "パウチ", "キャップ", "容器", "シュリンク",
                 "シール", "中栓", "台紙", "スポイド", "スポイト", "外装", "化粧箱",
                 "個箱", "内箱", "外箱", "6コ箱", "６コ箱")
_SHIZAI_SUFFIX_NORM = tuple(normalize(s) for s in SHIZAI_SUFFIX)

def is_shizai(name):
    if name is None:
        return True
    raw = str(name)
    if "残資材" in raw:
        return True
    n = normalize(raw)
    if n in _SHIZAI_NORM:
        return True
    return n.endswith(_SHIZAI_SUFFIX_NORM)  # 「…ラベル」等の資材行

# ================= 商品マスタ読み込み =================
def _sheets_from(fileobj, filename):
    """ファイルを [ (sheet_name, rows) ] に変換。rows は list[tuple(cells)]。"""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        data = fileobj.read()
        if isinstance(data, bytes):
            for enc in ("utf-8-sig", "utf-8", "cp932"):
                try:
                    text = data.decode(enc); break
                except UnicodeDecodeError:
                    text = data.decode("utf-8", "replace")
        else:
            text = data
        rows = list(csv.reader(io.StringIO(text)))
        return [("csv", rows)]
    wb = load_workbook(fileobj, data_only=True, read_only=True)
    out = []
    for ws in wb.worksheets:
        out.append((ws.title, list(ws.iter_rows(values_only=True))))
    return out

def _is_name_col(c):
    # 製品名/販売名に加え、原料タブの見出し「原料名」も名前列として認識する
    return ("製品名" in c) or ("販売名" in c) or ("原料名" in c) or ("品名" in c)

def _find_master_header(rows):
    """製品名(または原料名)・単価列を含むヘッダー行を探し、列マップを返す。"""
    for i, row in enumerate(rows[:30]):
        nc = [normalize(c) for c in row]
        has_name = any(_is_name_col(c) for c in nc)
        has_price = any("単価" in c for c in nc)
        if has_name and has_price:
            colmap = {}
            for j, c in enumerate(nc):
                if _is_name_col(c) and "製品名" not in colmap:
                    colmap["製品名"] = j
                if (("商品cd" in c) or ("商品コード" in c)) and "商品CD" not in colmap:
                    colmap["商品CD"] = j
                if "試作番号" in c and "試作番号" not in colmap:
                    colmap["試作番号"] = j
                if ("売上単価" in c or c == "単価") and "単価" not in colmap:
                    colmap["単価"] = j
                if "価格更新" in c and "価格更新" not in colmap:
                    colmap["価格更新"] = j
                if c == "ロット" and "ロット" not in colmap:
                    colmap["ロット"] = j
            if "製品名" in colmap and "単価" in colmap:
                return i, colmap
    return None, None

def load_master(fileobj, filename):
    """商品マスタを読み込み、正規化製品名 -> エントリ一覧 の索引を返す。"""
    index = {}
    n_entries = 0
    for sheet_name, rows in _sheets_from(fileobj, filename):
        hidx, colmap = _find_master_header(rows)
        if hidx is None:
            continue
        cn, cc = colmap["製品名"], colmap.get("商品CD")
        cs, cp = colmap.get("試作番号"), colmap["単価"]
        cu, cl = colmap.get("価格更新"), colmap.get("ロット")
        for row in rows[hidx + 1:]:
            if cn >= len(row):
                continue
            name = row[cn]
            if name is None or not str(name).strip():
                continue
            tanka = first_num(row[cp]) if cp < len(row) else None
            if tanka is None:
                continue
            entry = {
                "製品名": str(name).strip(),
                "商品CD": str(row[cc]).strip() if (cc is not None and cc < len(row) and row[cc] is not None) else "",
                "試作番号": str(row[cs]).strip() if (cs is not None and cs < len(row) and row[cs] is not None) else "",
                "単価": tanka,
                "価格更新": parse_date(row[cu]) if (cu is not None and cu < len(row)) else None,
                "ロット下限": first_num(row[cl]) if (cl is not None and cl < len(row)) else None,
                "シート": sheet_name,
            }
            index.setdefault(normalize(name), []).append(entry)
            n_entries += 1
    return index, n_entries

# ================= 単価選択 =================
def select_price(entries, qty):
    flags = []
    es = [e for e in entries if e["単価"] is not None]
    if not es:
        return None, ["単価が空"]
    thresholds = sorted({e["ロット下限"] for e in es if e["ロット下限"] is not None})
    if thresholds:
        le = [t for t in thresholds if t <= qty]
        if le:
            chosen_t = max(le)
        else:
            chosen_t = min(thresholds)
            flags.append("数量%d が最小ロット%d 未満→最小単価採用(要確認)" % (int(qty), int(chosen_t)))
        cands = [e for e in es if e["ロット下限"] == chosen_t]
    else:
        cands = es
    cds = {e["商品CD"] for e in cands if e["商品CD"]}
    if len(cds) > 1:
        flags.append("商品CD複数 %s →名前衝突(要確認)" % sorted(cds))
    chosen = max(cands, key=lambda e: e["価格更新"] or datetime.date(1900, 1, 1))
    return chosen, flags

# ================= 仮納品書 読み込み =================
def _colmap_kari(nc):
    cmap = {}
    for j, c in enumerate(nc):
        if "品名" in c and "品名" not in cmap:
            cmap["品名"] = j
        if "数量" in c and "数量" not in cmap:
            cmap["数量"] = j
        if ("lot" in c or "ロット" in c) and "lot" not in cmap:
            cmap["lot"] = j
        if "備考" in c and "備考" not in cmap:
            cmap["備考"] = j
    return cmap

def _is_red_font(cell):
    """赤文字セルか判定。黒=正本／赤=二重記載の控え とみなし、赤を除外するため。"""
    if cell is None:
        return False
    try:
        color = cell.font.color
    except Exception:
        return False
    if color is None:
        return False
    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str) and len(rgb) >= 6:
        h = rgb[-6:].upper()
        try:
            r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        except ValueError:
            return False
        return r >= 0x90 and g <= 0x70 and b <= 0x70  # 朱・赤・濃赤を赤系とみなす
    if getattr(color, "indexed", None) in (2, 10):  # 標準パレットの赤
        return True
    return False

_DEST_SKIP = ("仮納品書", "受領書", "日付", "御中", "下記", "サンブルーム", "大阪市",
              "工場", "品名", "数量", "合計", "ご担当", "申し上げ", "No.")
_DEST_SKIP_NORM = tuple(normalize(x) for x in _DEST_SKIP)
def _clean_dest(s):
    """発送先名から法人格・御中等を除く。例『株式会社ゼロ・インフィニティ』→『ゼロ・インフィニティ』"""
    s = unicodedata.normalize("NFKC", str(s or ""))
    s = re.sub(r"(株式会社|有限会社|合同会社|合資会社|㈱|㈲|\(株\)|（株）)", "", s)
    return s.replace("御中", "").strip("　 ").strip()

def _find_dest(rows, hidx, pcol):
    """ヘッダー前の行から発送先（宛先）を推定。先頭側の会社名らしい1件を返す。"""
    for row in rows[1:hidx]:
        for c in row:
            if c in (None, "") or isinstance(c, (datetime.date, datetime.datetime, int, float)):
                continue
            if parse_date(c):       # 日付っぽいセルは除外
                continue
            s = str(c).strip()
            sn = normalize(s)       # 全角スペース除去後で判定（"仮　納　品　書"対策）
            if not sn or any(k in sn for k in _DEST_SKIP_NORM):
                continue
            if len(sn) >= 3:
                return s
    return ""

def parse_karinouhin(fileobj):
    """仮納品書を解析。最初の「品名&数量」ブロックのみを読み、最初の「合計」で停止
    （続く受領書ブロックは読まない）。さらに行内が赤文字（黒との二重記載の控え）の
    明細はフォント色で判定してスキップし、二重計上を防ぐ。
    (items, excluded, kari_date) を返す。"""
    wb = load_workbook(fileobj, data_only=True)  # 色判定のため read_only にしない（仮納品書は小サイズ）
    items, excluded, date_votes = [], [], []
    for ws in wb.worksheets:
        cell_rows = list(ws.iter_rows())
        rows = [[c.value for c in r] for r in cell_rows]
        # シートごとの出荷日（メモ等の日付を拾わないようヘッダー付近に限定）
        sheet_votes = []
        for row in rows[:14]:
            if any("日付" in normalize(c) for c in row):
                for c in row:
                    d = parse_date(c)
                    if d:
                        sheet_votes.append(d)
        sheet_date = Counter(sheet_votes).most_common(1)[0][0] if sheet_votes else None
        if sheet_date:
            date_votes.append(sheet_date)
        sd_str = sheet_date.isoformat() if sheet_date else ""
        # 最初の「品名&数量」ヘッダー = 仮納品書の明細ブロック。
        hidx, cmap = None, None
        for i, row in enumerate(rows):
            nc = [normalize(c) for c in row]
            if any("品名" in c for c in nc) and any("数量" in c for c in nc):
                hidx, cmap = i, _colmap_kari(nc)
                break
        if hidx is None or "品名" not in cmap or "数量" not in cmap:
            continue
        pcol, qcol = cmap["品名"], cmap["数量"]
        lcol, bcol = cmap.get("lot"), cmap.get("備考")
        sheet_dest = _find_dest(rows, hidx, pcol)  # バルク出荷先の表示に使用
        zanzai = False  # 「○○ 残資材」ヘッダー以降はブロック全体が資材なので除外
        for ci in range(hidx + 1, len(cell_rows)):
            row = rows[ci]
            name = row[pcol] if pcol < len(row) else None
            nn = normalize(name)
            if "合計" in nn or "受領" in nn:
                break  # 仮納品書ブロックの終端で停止（受領書は読まない）
            if not nn:
                continue
            # 黒/赤の二重記載：赤文字の行は控え（黒の写し）なので除外
            crow = cell_rows[ci]
            name_cell = crow[pcol] if pcol < len(crow) else None
            qcell = row[qcol] if qcol < len(row) else ""
            if _is_red_font(name_cell):
                excluded.append({"製品名": str(name).strip(), "数量": qcell,
                                 "除外理由": "赤文字（二重記載の控え）", "シート": ws.title, "日付": sd_str})
                continue
            raw = str(name)
            # 「○○ 残資材」「ご支給原料」行＝資材/原料ブロックの開始。以降のブロック明細は全て除外。
            if "残資材" in raw or "支給" in raw:
                zanzai = True
                excluded.append({"製品名": raw.strip(), "数量": qcell,
                                 "除外理由": "残資材/支給ブロック(開始)", "シート": ws.title, "日付": sd_str})
                continue
            if zanzai:
                excluded.append({"製品名": raw.strip(), "数量": qcell,
                                 "除外理由": "残資材/支給ブロック", "シート": ws.title, "日付": sd_str})
                continue
            qty = first_num(qcell)
            lot = row[lcol] if (lcol is not None and lcol < len(row)) else None
            bikou = row[bcol] if (bcol is not None and bcol < len(row)) else None
            rec = {
                "製品名": raw.strip(),
                "数量": qty if qty is not None else "",
                "Lot": str(lot).strip() if lot else "",
                "備考": str(bikou).strip() if bikou else "",
                "シート": ws.title,
                "日付": sheet_date,           # 集計のグルーピングキー（date or None）
                "発送先": sheet_dest,         # バルク出荷先の表示用
            }
            if is_shizai(name):  # 個別名の資材（容器/ポンプ等）も念のため除外
                excluded.append({**rec, "日付": sd_str, "除外理由": "資材/残資材"})
                continue
            rec["数量"] = qty  # 数量が空でも除外しない（後段で「要確認」に回す）
            items.append(rec)
    kari_date = Counter(date_votes).most_common(1)[0][0] if date_votes else None
    return items, excluded, kari_date

# ================= 集計・突合 =================
def parse_case_terms(text):
    """備考のケース構成を (入数, 箱数) のリストへ。
    例 '300×3c/s、100×1c/s' -> [(300,3),(100,1)] / '1c/s'(バルク) -> [(None,1)]"""
    if not text:
        return []
    t = unicodedata.normalize("NFKC", str(text))
    pairs = []
    def _take(m):
        try:
            pairs.append((float(m.group(1).replace(",", "")), float(m.group(2).replace(",", ""))))
        except ValueError:
            pass
        return " "
    # 「入数×箱数」を先に拾って消す
    t2 = re.sub(r"(\d[\d,]*)\s*[x×*]\s*(\d[\d,]*)", _take, t)
    # 残った「N c/s」(×なし＝バルク等)は 入数なし×箱数
    for m in re.finditer(r"(\d[\d,]*)\s*[cＣ]\s*[/／]\s*[sＳ]", t2):
        try:
            pairs.append((None, float(m.group(1).replace(",", ""))))
        except ValueError:
            pass
    return pairs

def aggregate(items):
    """製品×ロットで数量を合算。ケースは入数ごとに箱数を合算（例 50×10 + 50×1×3 -> 50×13）。"""
    agg = {}
    for it in items:
        key = (normalize(it["製品名"]), it["Lot"])
        if key not in agg:
            agg[key] = {"製品名": it["製品名"], "Lot": it["Lot"], "数量": 0.0,
                        "cases": defaultdict(float), "数量欠落": False,
                        "発送先": it.get("発送先", "")}
        if not agg[key]["発送先"] and it.get("発送先"):
            agg[key]["発送先"] = it["発送先"]
        q = it["数量"]
        if q is None:
            agg[key]["数量欠落"] = True   # 数量が空の明細あり → 後段で要確認
        else:
            agg[key]["数量"] += q
        for nyusu, hako in parse_case_terms(it["備考"]):
            agg[key]["cases"][nyusu] += hako
    return list(agg.values())

# 詳細シートの列（メインの印刷シートとは別。金額は出さない）
DETAIL_COLS = ["日付", "製品名", "ロット", "出荷数", "ケース", "商品CD", "処方番号", "単価", "要確認", "メモ"]

def _num_out(q):
    """整数ならint、小数あり（バルクのkg等）ならそのまま小数で返す。"""
    q = float(q)
    return int(q) if q.is_integer() else round(q, 3)

def _case_lines(a):
    """入数の降順で (入数 or None, 箱数) のリスト。バルクは入数None。"""
    return [(n, h) for n, h in sorted(a["cases"].items(), key=lambda kv: -(kv[0] or 0)) if h]

def _cases_str(lines):
    """詳細表示用の文字列。例 [(12,16),(8,1)] -> '12×16、8×1' / [(None,1)] -> '1c/s'"""
    return "、".join(("%d×%d" % (int(n), int(h)) if n else "%dc/s" % int(h)) for n, h in lines)

# 製品名末尾の容量（10ml/980g/150ｍL 等）を分離。全角/半角の混在に対応した文字クラス。
_YORYO_UNIT = (r"(?:[mｍＭ][lLｌＬ]|[kＫｋ][gＧｇ]|㎏|㎖|[gＧｇ]|[lLｌＬℓ]|[cＣ][cＣ]|個|本|錠|包|枚)")
_YORYO_RE = re.compile(r"[\s　]*([0-9０-９][0-9０-９.．,，]*\s*" + _YORYO_UNIT + r")[\s　]*$")

def _split_name(name):
    """製品名から末尾の容量を分離。バルク(名前に『バルク』)は容量なし・bulk=True。
    返り値: (製品名_本体, 容量, bulk)"""
    raw = str(name).strip()
    if "バルク" in raw:
        main = re.sub(r"[\s　]*バルク[\s　]*$", "", raw).strip()
        return (main or raw), "", True
    m = _YORYO_RE.search(raw)
    if m:
        return raw[:m.start()].strip(), m.group(1).strip(), False
    return raw, "", False

def _norm_tokens(s):
    """NFKC+小文字でトークン分割（空白境界を残す）。全角スペースは半角化される。"""
    return unicodedata.normalize("NFKC", str(s)).lower().split()

def _fuzzy_master_lookup(name, master_index):
    """完全一致しない時、末尾の短いサフィックス(グレード記号 N/β 等)を外して再検索。
    例: 『Radical Sponge Ｎ』→『Radical Sponge』にマスタ一致。(entries, note) を返す。"""
    toks = _norm_tokens(name)
    if len(toks) >= 2 and len(toks[-1]) <= 2:
        loose = "".join(toks[:-1])  # 空白除去済み = normalize() 相当
        e = master_index.get(loose)
        if e:
            return e, "末尾「%s」を無視してマスタ一致" % toks[-1].upper()
    return [], ""

def build_records(agg, master_index):
    """製品ごとの records を返す（印刷シート・詳細シートの両方の元データ）。原料は除外。"""
    records = []
    for a in sorted(agg, key=lambda x: normalize(x["製品名"])):
        lines = _case_lines(a)
        name_main, yoryo, bulk = _split_name(a["製品名"])
        missing_all = a.get("数量欠落") and a["数量"] == 0  # 数量が全く取れていない
        note = ""
        if missing_all:
            chosen, flags = None, ["数量が空（要確認）"]
        else:
            entries = master_index.get(normalize(a["製品名"]), [])
            if not entries:  # 完全一致なし → 末尾サフィックスを外して再検索
                entries, note = _fuzzy_master_lookup(a["製品名"], master_index)
            chosen, flags = (None, ["単価リストに該当なし"]) if not entries else select_price(entries, a["数量"])
            flags = list(flags)
            if a.get("数量欠落"):
                flags.append("一部明細で数量が空（要確認）")
            if chosen and "原料" in (chosen.get("シート") or ""):  # マスタの原料タブ一致＝原料 → 載せない
                continue
        tanka = chosen["単価"] if chosen else None
        records.append({
            "製品名": a["製品名"], "name_main": name_main, "容量": yoryo, "bulk": bulk,
            "発送先": a.get("発送先", ""), "ロット": a["Lot"],
            "数量": (None if missing_all else a["数量"]),
            "cases": lines,
            "商品CD": chosen["商品CD"] if chosen else "",
            "処方番号": chosen["試作番号"] if chosen else "",
            "単価": int(tanka) if tanka else None,
            "要確認": bool(flags),
            "メモ": " / ".join([x for x in flags + ([note] if note else []) if x]),
        })
    return records

# ================= 出力(xlsx) =================
_FONT = "Yu Gothic"
_INVALID_SHEET = re.compile(r"[\[\]\:\*\?\/\\]")

def _safe_sheet_title(label, used):
    """Excelシート名に使える形へ（禁則文字除去・31文字制限・重複回避）。"""
    t = _INVALID_SHEET.sub("-", str(label or "出荷連絡表")).strip() or "出荷連絡表"
    t = t[:31]
    base, i = t, 2
    while t in used:
        suf = "(%d)" % i
        t = base[:31 - len(suf)] + suf
        i += 1
    used.add(t)
    return t

def _qty_disp(rec):
    """出荷数の表示。バルクは『22㎏』、それ以外は数量、空はブランク。"""
    if rec["数量"] is None:
        return ""
    v = _num_out(rec["数量"])
    return ("%s㎏" % v) if rec["bulk"] else v

# 実物テンプレ（事務所→工場）に合わせた列幅
_PRINT_W = {"A": 39.1, "B": 13.9, "C": 15.8, "D": 7.9, "E": 5.4, "F": 7.9, "G": 6.7}

def _write_print_sheet(ws, records, date_label):
    """出荷連絡.xlsx と同じ印刷フォーマット（A4縦・1製品2行・容量右詰・ケースは入数×箱数C/S）。"""
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cen = Alignment(horizontal="center", vertical="center")
    rgt = Alignment(horizontal="right", vertical="center")
    lft = Alignment(horizontal="left", vertical="center")
    # 印刷設定
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9   # A4
    ws.page_setup.scale = 90
    ws.page_margins.left = 0.59; ws.page_margins.right = 0.33
    ws.page_margins.top = 1.0; ws.page_margins.bottom = 0.45
    for col, w in _PRINT_W.items():
        ws.column_dimensions[col].width = w
    # 見出し
    ws.merge_cells("A1:B1")
    ws["A1"] = "出　荷　連　絡　表　"
    ws["A1"].font = Font(name=_FONT, bold=True, size=14)
    ws["A1"].alignment = cen
    ws["C1"] = date_label
    ws["C1"].font = Font(name=_FONT); ws["C1"].alignment = cen
    ws.merge_cells("A3:G3")
    ws["A3"] = "事務所（ＦＡＸ06-6453-3916）ｏｒ　honsha@sunbloom-cosme.co.jp　⇒工　場"
    ws["A3"].font = Font(name=_FONT); ws["A3"].alignment = cen
    for col, h in zip("ABCD", ["製品名", "ロット", "出荷数", "ケース数"]):
        c = ws["%s5" % col]; c.value = h
        c.font = Font(name=_FONT, bold=True); c.alignment = cen; c.border = border
    # 明細（1製品2行以上）
    r = 6
    for rec in records:
        nrows = max(2, len(rec["cases"]))
        nm = rec["name_main"] + ("　【要確認】" if rec["要確認"] else "")
        ws.cell(r, 1, nm).alignment = lft
        if rec["bulk"]:
            d = _clean_dest(rec["発送先"])
            sub = ("%sへバルク出荷" % d) if d else "バルク出荷"
        else:
            sub = rec["容量"]
        ws.cell(r + 1, 1, sub).alignment = rgt
        ws.cell(r, 2, rec["ロット"]).alignment = cen
        ws.cell(r, 3, _qty_disp(rec)).alignment = rgt
        for i in range(nrows):
            rr = r + i
            ws.cell(rr, 5, "×").alignment = cen
            ws.cell(rr, 7, "C/S").alignment = cen
            if i < len(rec["cases"]):
                nyu, hako = rec["cases"][i]
                if nyu:
                    ws.cell(rr, 4, int(nyu)).alignment = rgt
                ws.cell(rr, 6, int(hako)).alignment = rgt
        # 罫線＋フォント
        for rr in range(r, r + nrows):
            for cc in range(1, 8):
                cell = ws.cell(rr, cc)
                cell.border = border
                cell.font = Font(name=_FONT, size=10.5)
        # ロット・出荷数は製品の行数ぶん結合（実物同様）
        if nrows >= 2:
            ws.merge_cells(start_row=r, start_column=2, end_row=r + nrows - 1, end_column=2)
            ws.merge_cells(start_row=r, start_column=3, end_row=r + nrows - 1, end_column=3)
        r += nrows
    ws.print_area = "A1:G%d" % max(r - 1, 5)

_DETAIL_W = {"日付": 11, "製品名": 34, "ロット": 10, "出荷数": 9, "ケース": 16,
             "商品CD": 11, "処方番号": 16, "単価": 9, "要確認": 8, "メモ": 40}

def _write_detail_sheet(ws, dated_records):
    """確認用の詳細シート（商品CD・処方番号・単価・要確認・メモを含む）。"""
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cols = DETAIL_COLS
    ws.merge_cells("A1:%s1" % ws.cell(row=1, column=len(cols)).column_letter)
    ws["A1"] = "出荷連絡表（詳細・確認用）"
    ws["A1"].font = Font(name=_FONT, bold=True, size=12)
    for j, h in enumerate(cols, 1):
        c = ws.cell(2, j, h); c.font = Font(name=_FONT, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border
    i = 3
    for date_label, rec in dated_records:
        vals = {
            "日付": date_label, "製品名": rec["製品名"], "ロット": rec["ロット"],
            "出荷数": ("" if rec["数量"] is None else _num_out(rec["数量"])),
            "ケース": _cases_str(rec["cases"]),
            "商品CD": rec["商品CD"], "処方番号": rec["処方番号"], "単価": rec["単価"],
            "要確認": ("要確認" if rec["要確認"] else ""), "メモ": rec["メモ"],
        }
        for j, k in enumerate(cols, 1):
            c = ws.cell(i, j, vals.get(k, "")); c.font = Font(name=_FONT, size=10.5); c.border = border
            if k in ("単価", "出荷数"):
                c.number_format = "#,##0.###"
                c.alignment = Alignment(horizontal="right", vertical="top")
            else:
                c.alignment = Alignment(wrap_text=True, vertical="top")
        i += 1
    for j, k in enumerate(cols, 1):
        ws.column_dimensions[ws.cell(row=2, column=j).column_letter].width = _DETAIL_W.get(k, 12)
    ws.freeze_panes = "A3"

def to_workbook(groups, *_ignore, **_kw):
    """groups = [{"日付","records":[...]}, ...] を、出荷日ごとの印刷シート＋詳細シート1枚に。"""
    wb = Workbook(); wb.remove(wb.active)
    used = set()
    for g in groups:
        ws = wb.create_sheet(_safe_sheet_title(g["日付"], used))
        _write_print_sheet(ws, g.get("records", []), g["日付"])
    dws = wb.create_sheet(_safe_sheet_title("詳細", used))
    _write_detail_sheet(dws, [(g["日付"], r) for g in groups for r in g.get("records", [])])
    if not wb.worksheets:
        wb.create_sheet("出荷連絡表")
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio

# ================= 一括処理 =================
def generate(karinouhin_fileobjs, master_fileobj, master_filename, date_label=""):
    """複数の仮納品書（各々が複数シートでも可）を受け取り、出荷日ごとにグルーピングして
    「1日付＝1シート」の出荷連絡表グループ群を返す。
    返り値: (groups, stats, debug)
      groups = [ {"日付": "2026-06-19", "rows": [...], "colorder": [...], "明細数": n, "要確認": n}, ... ]
    """
    if not isinstance(karinouhin_fileobjs, (list, tuple)):
        karinouhin_fileobjs = [karinouhin_fileobjs]
    master_index, n = load_master(master_fileobj, master_filename)
    items, excluded, per_file = [], [], []
    for f in karinouhin_fileobjs:
        it, ex, kd = parse_karinouhin(f)
        items += it
        excluded += ex
        per_file.append({"ファイル": getattr(f, "name", "(不明)"),
                         "明細(資材除外後)": len(it), "除外": len(ex),
                         "日付": kd.isoformat() if kd else "(取得できず)"})
    # 出荷日でグルーピング（手入力があれば全件をそのラベル1シートへ）
    all_dates = [it["日付"] for it in items if it.get("日付")]
    overall = Counter(all_dates).most_common(1)[0][0] if all_dates else None
    override = (date_label or "").strip()
    today = datetime.date.today().isoformat()
    groups_map = defaultdict(list)
    if override:
        groups_map[override] = list(items)
    else:
        for it in items:
            d = it.get("日付") or overall          # 日付が取れないシートは代表日へ
            groups_map[d.isoformat() if d else today].append(it)
    groups = []
    for label in sorted(groups_map.keys()):
        recs = build_records(aggregate(groups_map[label]), master_index)
        groups.append({"日付": label, "records": recs,
                       "明細数": len(groups_map[label]),
                       "要確認": sum(1 for r in recs if r["要確認"])})
    if not groups:  # 明細ゼロでも空の1シートは作る
        groups.append({"日付": override or (overall.isoformat() if overall else today),
                       "records": [], "明細数": 0, "要確認": 0})
    stats = {
        "マスタ商品数": n,
        "仮納品書 明細数(資材除外後)": len(items),
        "出荷連絡表 行数": sum(len(g["records"]) for g in groups),
        "要確認 行数": sum(g["要確認"] for g in groups),
        "シート数": len(groups),
        "日付一覧": [g["日付"] for g in groups],
        "日付自動取得": bool(all_dates) and not override,
        "ファイル数": len(karinouhin_fileobjs),
    }
    debug = {"items": items, "excluded": excluded, "per_file": per_file}
    return groups, stats, debug

if __name__ == "__main__":
    import sys
    # 使い方: python engine.py <仮納品書...> <master> [out(出荷連絡表*.xlsx)]
    args = sys.argv[1:]
    out = "出力_出荷連絡表.xlsx"
    if args and args[-1].lower().endswith(".xlsx") and "出荷連絡表" in args[-1]:
        out = args.pop()
    master = args.pop()
    karis = args
    kfs = [open(k, "rb") for k in karis]
    with open(master, "rb") as mf:
        groups, stats, debug = generate(kfs, mf, master)
    for f in kfs:
        f.close()
    print("マスタ商品数:", stats["マスタ商品数"], "/ 明細:", stats["仮納品書 明細数(資材除外後)"],
          "/ 除外:", len(debug["excluded"]), "/ シート数:", stats["シート数"], "/ 日付:", stats["日付一覧"])
    for g in groups:
        print("=" * 90, "\n■ 出荷日:", g["日付"], "（", len(g["records"]), "製品 ）")
        for r in g["records"]:
            print("%-24s 容量:%-7s Lot:%-5s 数:%7s ケース:%-14s 単価:%s %s%s" % (
                r["name_main"][:24], r["容量"], r["ロット"],
                ("" if r["数量"] is None else str(_num_out(r["数量"]))),
                _cases_str(r["cases"]), str(r["単価"]),
                ("[バルク→%s]" % _clean_dest(r["発送先"])) if r["bulk"] else "",
                ("[要確認 " + r["メモ"] + "]") if r["要確認"] else ""))
    with open(out, "wb") as f:
        f.write(to_workbook(groups).read())
    print("-" * 90); print("saved:", out)
